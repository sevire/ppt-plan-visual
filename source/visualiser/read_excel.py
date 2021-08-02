import openpyxl

def read_excel(excel_path, sheet_name, skiprows=0):
    """
    Meant to be a replacement for using Pandas in plan visualiser so trying to keep as simple as possible for now.

    - Start at row implied by skiprows
    - Treat first row as headings for columns
    - First blank heading denotes last column of data
    - Last row of data defined by first row with blanks in all columns
    - Store the data in a dictionary with headings as key
    - Store each column as an array under the key of column heading
    - Allow iteration through rows returning dict for each row under column headings

    :param excel_path:
    :param sheet_name:
    :param skiprows:
    :return:
    """
    wb_obj = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb_obj[sheet_name]
    start_row = 1+skiprows

    table = {}
    headings = get_headers(sheet, start_row)
    table = {heading:[] for heading in headings}

    row_confirmed = True
    read_row_num = start_row
    while row_confirmed:
        row_confirmed = read_row(read_row_num, sheet, headings, table)
        read_row_num += 1

    iterable_by_row = iterrows(table)
    return iterable_by_row


def read_row(table_row_num, sheet, headings, table, skiprows=0):
    maybe_row = {}
    row_confirmed = False
    for col, heading in enumerate(headings):
        cell_value = sheet.cell(table_row_num + skiprows + 1, col + 1).value
        maybe_row[heading] = cell_value
        if cell_value is not None:
            row_confirmed = True
    if row_confirmed:
        for heading in headings:
            table[heading].append(maybe_row[heading])
    return row_confirmed


def get_headers(sheet, start_row):
    headings = []  # Need to store headings in column order so list not dict
    blank = False
    row = start_row
    column = 1
    while not blank:
        maybe_heading = sheet.cell(row, column).value
        if maybe_heading is not None:
            headings.append(maybe_heading)
            column += 1
        else:
            blank = True
    return headings


def iterrows(excel_table):
    # Work out num rows by checking any of the column arrays
    num_rows = len(next(iter(excel_table.values())))
    return [{heading:excel_table[heading][row] for heading in excel_table} for row in range(num_rows)]
